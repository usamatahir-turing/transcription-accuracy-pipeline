"""Generate Excel transcription accuracy reports from per-conversation metrics.json files.

Reads ``Conversations/<batch>/<conversation>/metrics.json`` and writes one
``.xlsx`` workbook with:

  - one worksheet per batch (per-conversation + by-language summary)
  - a final **All Batches** worksheet combining every batch
  - a **top_erroneous_segments** worksheet (worst segments per speaker from
    ``SPK*_top_errors.json``, produced by ``rank_error_segments.py``)
  - a **Definitions** worksheet explaining each metric column
  - a **Reference** worksheet with baseline reference numbers
  - per-batch **DetER** tables (column Y→) from ``deter.json`` when present

Human-annotated transcript metrics are compared to the independent vendor
baseline reference. Delta columns show *annotated transcript − baseline* in
percentage points; lower error rates are better, so negative deltas are green
and positive deltas are amber/red.

Usage
-----
    .\\.venv\\Scripts\\python.exe generate_report.py
    .\\.venv\\Scripts\\python.exe generate_report.py --output reports/transcription_accuracy_metrics.xlsx
    .\\.venv\\Scripts\\python.exe generate_report.py --batch delivery_batch_06092026
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from workflow_common import add_scope_args, resolve_conversation_dirs

# Baseline reference numbers (independent vendor corpus, identical pipeline).
REFERENCE = {
    "AR": {"wer": 47.04, "cer": 20.28, "wcmr": 63.22, "gt3_pct": 28.49},
    "GR": {"wer": 15.62, "cer": 10.24, "wcmr": 38.92, "gt3_pct": 9.57},
    "EN": {"wer": 8.51, "cer": 6.40, "wcmr": 28.22, "gt3_pct": 3.33},
    "ES": {"wer": 14.37, "cer": 10.73, "wcmr": 36.01, "gt3_pct": 6.50},
    "FR": {"wer": 17.20, "cer": 12.97, "wcmr": 51.12, "gt3_pct": 25.78},
    "IT": {"wer": 11.97, "cer": 7.90, "wcmr": 40.61, "gt3_pct": 9.14},
    "JA": {"wer": 71.83, "cer": 11.77, "wcmr": 43.74, "gt3_pct": 16.16},
    "KO": {"wer": 18.21, "cer": 10.13, "wcmr": 28.27, "gt3_pct": 2.51},
    "PT": {"wer": 12.31, "cer": 6.92, "wcmr": 46.02, "gt3_pct": 8.62},
    "RU": {"wer": 16.80, "cer": 11.13, "wcmr": 54.66, "gt3_pct": 6.39},
    "ALL": {"wer": 19.38, "cer": 10.73, "wcmr": 42.42, "gt3_pct": 11.37},
}

# Transcription tables: language summary A–K, gap L, conversations M→ (width varies).
LANG_COL = 1
GAP_COL = 12
CONV_COL = 13

# DetER block starts two columns after the per-conversation table (one gap column).
# Positions are computed per sheet — see ``_deter_block_cols``.

# Per-channel DetER pass threshold (matches chsep_audio_qa DER_CH_MAX).
DETER_PASS_THRESHOLD_PCT = 10.0

# Blank columns after the last data column so horizontal scroll has breathing room.
TRAILING_PAD_COLS = 4
TRAILING_PAD_WIDTH = 14

# DetER tables reserve up to SPK08 speaker columns (layout scales with batch scope).
MAX_DETER_SPEAKER_COLS = 8

# Slim tables: core metrics + colored delta columns (vs baseline reference).
CONV_HEADERS = [
    ("Conversation", "session_id", None, False),
    ("Language", "language", None, False),
    ("Scored", "n_scored", "#,##0", False),
    ("WER %", "wer_pct", "0.00", False),
    ("Δ WER", "d_wer", "+0.0;-0.0", True),
    ("CER %", "cer_pct", "0.00", False),
    ("Δ CER", "d_cer", "+0.0;-0.0", True),
    ("WCMR %", "wcmr_pct", "0.00", False),
    ("Δ WCMR", "d_wcmr", "+0.0;-0.0", True),
    ("|m-n|>3 %", "gt3_pct", "0.00", False),
    ("Δ >3", "d_gt3", "+0.0;-0.0", True),
]

LANG_HEADERS = [
    ("Language", "language", None, False),
    ("Conv.", "n_conversations", "#,##0", False),
    ("Scored", "n_scored", "#,##0", False),
    ("WER %", "wer_pct", "0.00", False),
    ("Δ WER", "d_wer", "+0.0;-0.0", True),
    ("CER %", "cer_pct", "0.00", False),
    ("Δ CER", "d_cer", "+0.0;-0.0", True),
    ("WCMR %", "wcmr_pct", "0.00", False),
    ("Δ WCMR", "d_wcmr", "+0.0;-0.0", True),
    ("|m-n|>3 %", "gt3_pct", "0.00", False),
    ("Δ >3", "d_gt3", "+0.0;-0.0", True),
]

REF_HEADERS = [
    ("Language", "language", None, False),
    ("Scored (ref)", "n_scored", "#,##0", False),
    ("WER %", "wer", "0.00", False),
    ("CER %", "cer", "0.00", False),
    ("WCMR %", "wcmr", "0.00", False),
    ("|m-n|>3 %", "gt3_pct", "0.00", False),
]

REF_DETER_COL = 8
REF_DETER_HEADERS = [
    ("DetER pass ≤ %", "deter_pass_max", "0.00", False),
]

TOP_ERRORS_HEADERS = [
    ("Batch", "batch", None, False),
    ("Conversation", "session_id", None, False),
    ("Speaker", "speaker", None, False),
    ("Language", "language", None, False),
    ("Rank metric", "rank_metric", None, False),
    ("Rank", "rank", "#,##0", False),
    ("Idx", "idx", "#,##0", False),
    ("Start (s)", "start", "0.00", False),
    ("End (s)", "end", "0.00", False),
    ("Ref units", "ref_units", "#,##0", False),
    ("Hyp units", "hyp_units", "#,##0", False),
    ("Errors", "errors", "#,##0", False),
    ("Error rate %", "error_rate_pct", "0.00", False),
    ("Substitutions", "substitutions", "#,##0", False),
    ("Deletions", "deletions", "#,##0", False),
    ("Insertions", "insertions", "#,##0", False),
    ("Ref (norm)", "ref_norm", None, False),
    ("Hyp (norm)", "hyp_norm", None, False),
    ("Ref (raw)", "ref_raw", None, False),
    ("Hyp (raw)", "hyp_raw", None, False),
]

TOP_ERRORS_TEXT_KEYS = {
    "batch", "session_id", "speaker", "language", "rank_metric",
    "ref_norm", "hyp_norm", "ref_raw", "hyp_raw",
}
TOP_ERRORS_WRAP_KEYS = {"ref_norm", "hyp_norm", "ref_raw", "hyp_raw"}

REF_N_SCORED = {
    "AR": 2352, "GR": 2058, "EN": 3157, "ES": 2799, "FR": 3224,
    "IT": 2231, "JA": 5290, "KO": 4336, "PT": 3931, "RU": 3037, "ALL": 32415,
}

METRIC_DEFINITIONS = [
    (
        "Scored",
        "Number of speech segments included in scoring. Segments where the "
        "human-annotated reference has no real content after normalization "
        "(empty, filler-only, or backchannel-only) are excluded.",
    ),
    (
        "WER %",
        "Word Error Rate — (substitutions + deletions + insertions) ÷ reference "
        "word count, micro-averaged across scored segments. Compares human-"
        "annotated transcript vs Qwen3-ASR. Whitespace-delimited tokens after "
        "normalization. Lower is better.",
    ),
    (
        "Δ WER",
        "Human-annotated transcript WER % minus the baseline reference WER % "
        "for that language (percentage points). Negative = better than baseline.",
    ),
    (
        "CER %",
        "Character Error Rate — same as WER but on characters with whitespace "
        "removed before comparison. Human annotation vs Qwen3-ASR, micro-averaged. "
        "Lower is better. Preferred metric for Japanese (and useful when WER/WCMR "
        "are unreliable).",
    ),
    (
        "Δ CER",
        "Human-annotated transcript CER % minus the baseline reference CER % "
        "for that language (percentage points). Negative = better than baseline.",
    ),
    (
        "WCMR %",
        "Word-Count Mismatch Rate — share of scored segments where the human "
        "annotation and Qwen3-ASR hypothesis disagree on word count "
        "(len(ref.split()) ≠ len(hyp.split())). Flags text↔segment alignment "
        "issues. Lower is better.",
    ),
    (
        "Δ WCMR",
        "Human-annotated transcript WCMR % minus the baseline reference WCMR % "
        "for that language (percentage points). Negative = better than baseline.",
    ),
    (
        "|m-n|>3 %",
        "Share of scored segments where the absolute word-count gap between "
        "human annotation (m) and Qwen3-ASR output (n) exceeds 3. Large gaps "
        "often indicate segments whose text does not line up with the audio. "
        "Lower is better.",
    ),
    (
        "Δ >3",
        "Human-annotated transcript |m-n|>3 % minus the baseline reference value "
        "for that language (percentage points). Negative = better than baseline.",
    ),
    (
        "DetER % (SPKxx)",
        "Detection error rate per speaker channel — Sortformer ∪ Silero SAD vs "
        "speech-only seglst reference (collar 0.25 s). Missed speech + false "
        "alarm only; no speaker confusion. Lower is better.",
    ),
    (
        "DetER Pass",
        "Per conversation: PASS when every speaker channel is ≤ 10% DetER "
        "(DER_CH_MAX); FAIL otherwise.",
    ),
    (
        "Pass (# fail conv.)",
        "DetER-by-language summary only: FAIL (n) means n conversations in "
        "that language failed session pass (≥1 channel above 10%). PASS when "
        "all conversations in the language passed.",
    ),
    (
        "DetER pass ≤ %",
        "Per-channel DetER pass threshold (DER_CH_MAX) from chsep_audio_qa — "
        "10% for isolated speaker channels (language-agnostic).",
    ),
]

TOP_ERRORS_DEFINITIONS = [
    (
        "Rank metric",
        "How segments were ranked within each speaker file: CER for Japanese "
        "and Korean (character errors), WER for all other languages (word errors).",
    ),
    (
        "Rank",
        "1 = worst segment for that speaker (highest absolute error count).",
    ),
    (
        "Ref / Hyp units",
        "Reference and hypothesis unit counts for that segment (words, or "
        "characters for Japanese and Korean).",
    ),
    (
        "Errors",
        "Substitutions + deletions + insertions for that segment (S + D + I).",
    ),
    (
        "Error rate %",
        "Segment-level error rate: errors ÷ ref units × 100. Can exceed 100% "
        "when insertions dominate (e.g. vocalized Arabic Qwen output).",
    ),
    (
        "Ref / Hyp (norm)",
        "Normalized text used for scoring (after BasicTextNormalizer and "
        "filler stripping).",
    ),
    (
        "Ref / Hyp (raw)",
        "Original transcript strings before normalization (from the _norm.jsonl "
        "text field).",
    ),
]


def _pct(num: int, den: int) -> float | None:
    return round(100.0 * num / den, 2) if den else None


def _delta(actual: float | None, baseline: float | None) -> float | None:
    if actual is None or baseline is None:
        return None
    return round(actual - baseline, 2)


def _baseline_for(lang: str) -> dict:
    if lang == "ALL":
        return REFERENCE["ALL"]
    return REFERENCE.get(lang, {})


def enrich_with_deltas(record: dict) -> dict:
    """Add delta fields vs baseline reference for the row's language."""
    ref = _baseline_for(record.get("language", ""))
    out = dict(record)
    out["d_wer"] = _delta(record.get("wer_pct"), ref.get("wer"))
    out["d_cer"] = _delta(record.get("cer_pct"), ref.get("cer"))
    out["d_wcmr"] = _delta(record.get("wcmr_pct"), ref.get("wcmr"))
    out["d_gt3"] = _delta(record.get("gt3_pct"), ref.get("gt3_pct"))
    return out


def load_metrics(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def metrics_to_record(batch: str, data: dict) -> dict:
    """Flatten conversation-level ``total`` block into a row dict (+ raw counts)."""
    t = data["total"]
    wer = t.get("wer") or {}
    cer = t.get("cer") or {}
    wcmr = t.get("wcmr") or {}
    buckets = wcmr.get("buckets") or {}

    def _b(key: str) -> dict:
        return buckets.get(key) or {}

    record = {
        "batch": batch,
        "session_id": data["session_id"],
        "language": data.get("language") or "",
        "n_segments_total": t.get("n_segments_total", 0),
        "n_scored": t.get("n_scored", 0),
        "wer_pct": wer.get("pct"),
        "cer_pct": cer.get("pct"),
        "wcmr_pct": wcmr.get("pct"),
        "n_mismatch": wcmr.get("n_mismatch", 0),
        "gt3_n": _b("gt3").get("n", 0),
        "gt3_pct": _b("gt3").get("pct"),
        "ref_words": wer.get("ref_words", 0),
        "wer_s": wer.get("substitutions", 0),
        "wer_d": wer.get("deletions", 0),
        "wer_i": wer.get("insertions", 0),
        "ref_chars": cer.get("ref_chars", 0),
        "cer_s": cer.get("substitutions", 0),
        "cer_d": cer.get("deletions", 0),
        "cer_i": cer.get("insertions", 0),
    }
    return enrich_with_deltas(record)


def aggregate_records(records: list[dict], label: str) -> dict:
    """Micro-average a list of conversation records into one summary row."""
    n_scored = sum(r["n_scored"] for r in records)
    ref_words = sum(r["ref_words"] for r in records)
    ref_chars = sum(r["ref_chars"] for r in records)
    wer_err = sum(r["wer_s"] + r["wer_d"] + r["wer_i"] for r in records)
    cer_err = sum(r["cer_s"] + r["cer_d"] + r["cer_i"] for r in records)
    n_mismatch = sum(r["n_mismatch"] for r in records)
    gt3_n = sum(r.get("gt3_n", 0) for r in records)

    row = {
        "language": label,
        "n_conversations": len(records),
        "n_scored": n_scored,
        "wer_pct": _pct(wer_err, ref_words),
        "cer_pct": _pct(cer_err, ref_chars),
        "wcmr_pct": _pct(n_mismatch, n_scored),
        "n_mismatch": n_mismatch,
        "gt3_pct": _pct(gt3_n, n_scored),
    }
    return enrich_with_deltas(row)


def language_summary(records: list[dict]) -> list[dict]:
    by_lang: dict[str, list[dict]] = defaultdict(list)
    for r in records:
        by_lang[r["language"]].append(r)
    rows = [aggregate_records(by_lang[lang], lang) for lang in sorted(by_lang)]
    if rows:
        rows.append(aggregate_records(records, "ALL"))
    return rows


def derive_language(session_id: str) -> str:
    """Language code from session name, e.g. NV-KO-SS03-CONVO07 → KO."""
    parts = session_id.split("-")
    return parts[1] if len(parts) >= 2 else ""


def speaker_columns(records: list[dict]) -> list[str]:
    """SPK01…SPKnn for DetER tables (contiguous, up to ``MAX_DETER_SPEAKER_COLS``)."""
    max_n = 0
    for record in records:
        for key in record:
            if key.startswith("SPK") and key[3:].isdigit():
                max_n = max(max_n, int(key[3:]))
    if max_n == 0:
        return []
    n = min(max_n, MAX_DETER_SPEAKER_COLS)
    return [f"SPK{i:02d}" for i in range(1, n + 1)]


def load_deter(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def deter_to_record(batch: str, data: dict, language: str = "") -> dict:
    """Flatten ``deter.json`` into a row with one column per speaker channel."""
    lang = language or derive_language(data.get("session_id", ""))
    conv = data.get("conversation") or {}
    speakers = data.get("speakers") or {}
    record: dict = {
        "batch": batch,
        "session_id": data.get("session_id", ""),
        "language": lang,
        "deter_pass": conv.get("pass"),
        "mean_deter_pct": conv.get("mean_deter_pct"),
    }
    for spk, info in speakers.items():
        deter = (info or {}).get("deter") or {}
        record[spk] = deter.get("error_rate_pct")
    record["deter_pass_label"] = (
        "PASS" if record["deter_pass"] else "FAIL"
        if record["deter_pass"] is not None else "—"
    )
    return record


def discover_deter_records(root: Path, batch: str | None) -> dict[str, list[dict]]:
    """Return {batch_name: [DetER rows]} for conversations with deter.json."""
    batches: dict[str, list[dict]] = defaultdict(list)
    for conv_dir in resolve_conversation_dirs(root, batch, None):
        deter_path = conv_dir / "deter.json"
        if not deter_path.is_file():
            continue
        batch_name = conv_dir.parent.name
        metrics_path = conv_dir / "metrics.json"
        language = ""
        if metrics_path.is_file():
            language = load_metrics(metrics_path).get("language") or ""
        data = load_deter(deter_path)
        batches[batch_name].append(
            deter_to_record(batch_name, data, language=language))
    for batch_name in batches:
        batches[batch_name].sort(key=lambda r: r["session_id"])
    return dict(sorted(batches.items()))


def _avg(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def deter_language_summary(
    records: list[dict], speaker_cols: list[str],
) -> list[dict]:
    """Average per-speaker DetER within each language (+ ALL row)."""
    by_lang: dict[str, list[dict]] = defaultdict(list)
    for record in records:
        by_lang[record["language"]].append(record)

    def _summarize(group: list[dict], label: str) -> dict:
        row: dict = {
            "language": label,
            "n_conversations": len(group),
        }
        n_fail = sum(1 for r in group if r.get("deter_pass") is False)
        if n_fail:
            row["deter_pass_label"] = f"FAIL ({n_fail})"
        else:
            row["deter_pass_label"] = "PASS" if group else "—"
        for spk in speaker_cols:
            vals = [r[spk] for r in group if r.get(spk) is not None]
            row[spk] = _avg(vals)
        return row

    rows = [_summarize(by_lang[lang], lang) for lang in sorted(by_lang)]
    if rows:
        rows.append(_summarize(records, "ALL"))
    return rows


def build_deter_lang_headers(speaker_cols: list[str]) -> list[tuple]:
    headers: list[tuple] = [
        ("Language", "language", None, False),
        ("Conv.", "n_conversations", "#,##0", False),
    ]
    for spk in speaker_cols:
        headers.append((spk, spk, "0.00", "deter"))
    headers.append(("Pass (# fail conv.)", "deter_pass_label", None, False))
    return headers


def build_deter_conv_headers(
    speaker_cols: list[str], *, include_batch: bool = False,
) -> list[tuple]:
    headers: list[tuple] = []
    if include_batch:
        headers.append(("Batch", "batch", None, False))
    headers += [
        ("Conversation", "session_id", None, False),
        ("Language", "language", None, False),
    ]
    for spk in speaker_cols:
        headers.append((spk, spk, "0.00", "deter"))
    headers.append(("Pass", "deter_pass_label", None, False))
    return headers


def discover_records(root: Path, batch: str | None) -> dict[str, list[dict]]:
    """Return {batch_name: [conversation records]} for conversations with metrics.json."""
    batches: dict[str, list[dict]] = defaultdict(list)
    for conv_dir in resolve_conversation_dirs(root, batch, None):
        metrics_path = conv_dir / "metrics.json"
        if not metrics_path.is_file():
            continue
        batch_name = conv_dir.parent.name
        data = load_metrics(metrics_path)
        batches[batch_name].append(metrics_to_record(batch_name, data))
    for batch_name in batches:
        batches[batch_name].sort(key=lambda r: r["session_id"])
    return dict(sorted(batches.items()))


def discover_top_error_segments(root: Path, batch: str | None) -> list[dict]:
    """Flatten every ``SPK*_top_errors.json`` in scope into one row per segment."""
    rows: list[dict] = []
    for conv_dir in resolve_conversation_dirs(root, batch, None):
        batch_name = conv_dir.parent.name
        for path in sorted(conv_dir.glob("SPK*_top_errors.json")):
            data = json.loads(path.read_text(encoding="utf-8"))
            session_id = data.get("session_id", conv_dir.name)
            speaker = data.get("speaker", path.stem.replace("_top_errors", ""))
            language = data.get("language", "")
            rank_metric = data.get("rank_metric", "")
            for seg in data.get("segments", []):
                rows.append({
                    "batch": batch_name,
                    "session_id": session_id,
                    "speaker": speaker,
                    "language": language,
                    "rank_metric": rank_metric,
                    "rank": seg.get("rank"),
                    "idx": seg.get("idx"),
                    "start": seg.get("start"),
                    "end": seg.get("end"),
                    "ref_units": seg.get("ref_units"),
                    "hyp_units": seg.get("hyp_units"),
                    "errors": seg.get("errors"),
                    "error_rate_pct": seg.get("error_rate_pct"),
                    "substitutions": seg.get("substitutions"),
                    "deletions": seg.get("deletions"),
                    "insertions": seg.get("insertions"),
                    "ref_norm": seg.get("ref_norm", ""),
                    "hyp_norm": seg.get("hyp_norm", ""),
                    "ref_raw": seg.get("ref_raw", ""),
                    "hyp_raw": seg.get("hyp_raw", ""),
                })
    rows.sort(key=lambda r: (
        r["batch"], r["session_id"], r["speaker"], r.get("rank") or 0,
    ))
    return rows


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

CLR_TITLE = "1F3864"
CLR_HEADER = "2F5496"
CLR_SECTION = "D6DCE4"
CLR_ALT = "F2F2F2"
CLR_TOTAL = "E0E0E0"

# Delta color scale (lower error = better → negative Δ is good).
CLR_MUCH_BETTER = "A9D18E"   # ≤ −5 pp
CLR_BETTER = "C6EFCE"        # −5 to −2
CLR_NEAR = "FFFFFF"          # −2 to +2
CLR_WORSE = "FFEB9C"         # +2 to +5
CLR_BAD = "FFC7CE"           # +5 to +10
CLR_MUCH_WORSE = "FF9999"    # > +10

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True, color="595959")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=CLR_TITLE)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=10)
FONT_TOTAL = Font(name="Calibri", size=10, bold=True)
FONT_LEGEND = Font(name="Calibri", size=9, italic=True, color="595959")
FONT_DELTA_GOOD = Font(name="Calibri", size=10, color="375623")
FONT_DELTA_OK = Font(name="Calibri", size=10, color="404040")
FONT_DELTA_WARN = Font(name="Calibri", size=10, color="9C6500")
FONT_DELTA_BAD = Font(name="Calibri", size=10, color="9C0006")

FILL_TITLE = PatternFill("solid", fgColor=CLR_TITLE)
FILL_HEADER = PatternFill("solid", fgColor=CLR_HEADER)
FILL_SECTION = PatternFill("solid", fgColor=CLR_SECTION)
FILL_ALT = PatternFill("solid", fgColor=CLR_ALT)
FILL_TOTAL = PatternFill("solid", fgColor=CLR_TOTAL)
FILL_DETER_PASS = PatternFill("solid", fgColor=CLR_BETTER)
FILL_DETER_FAIL = PatternFill("solid", fgColor=CLR_BAD)
FILL_DETER_PASS_LABEL = PatternFill("solid", fgColor=CLR_BETTER)
FILL_DETER_FAIL_LABEL = PatternFill("solid", fgColor=CLR_BAD)
FILL_DELTA = {
    "much_better": PatternFill("solid", fgColor=CLR_MUCH_BETTER),
    "better": PatternFill("solid", fgColor=CLR_BETTER),
    "near": PatternFill("solid", fgColor=CLR_NEAR),
    "worse": PatternFill("solid", fgColor=CLR_WORSE),
    "bad": PatternFill("solid", fgColor=CLR_BAD),
    "much_worse": PatternFill("solid", fgColor=CLR_MUCH_WORSE),
}

THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _delta_band(delta: float | None) -> str | None:
    if delta is None:
        return None
    if delta <= -5:
        return "much_better"
    if delta <= -2:
        return "better"
    if delta <= 2:
        return "near"
    if delta <= 5:
        return "worse"
    if delta <= 10:
        return "bad"
    return "much_worse"


def _delta_font(band: str | None) -> Font:
    if band in ("much_better", "better"):
        return FONT_DELTA_GOOD
    if band in ("bad", "much_worse"):
        return FONT_DELTA_BAD
    if band == "worse":
        return FONT_DELTA_WARN
    return FONT_DELTA_OK


def _safe_sheet_name(name: str) -> str:
    bad = set(r'\/*?:[]')
    cleaned = "".join(c if c not in bad else "_" for c in name)
    return cleaned[:31]


def _write_title_block(ws, title: str, subtitle: str, n_cols: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_LEFT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = FONT_SUBTITLE
    c2.alignment = ALIGN_LEFT
    return 3


def _write_reading_guide(ws, row: int, merge_end: int) -> int:
    """Legend + notes at the top, before any data tables."""
    items = [
        ("≤ −5", "much_better"),
        ("−5 to −2", "better"),
        ("±2", "near"),
        ("+2 to +5", "worse"),
        ("+5 to +10", "bad"),
        ("> +10", "much_worse"),
    ]
    label = ws.cell(row=row, column=LANG_COL, value="Δ color key (pp vs baseline):")
    label.font = FONT_LEGEND
    label.alignment = ALIGN_LEFT
    for i, (text, band) in enumerate(items, start=2):
        cell = ws.cell(row=row, column=LANG_COL + i - 1, value=text)
        cell.font = FONT_BODY
        cell.fill = FILL_DELTA[band]
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    row += 1
    callout_font = InlineFont(rFont="Calibri", sz=8, i=True, color="595959")
    note = CellRichText(
        "Δ = human-annotated transcript minus baseline reference (percentage points; "
        "lower error is better). See Definitions and Reference tabs for details.",
        "\n",
        TextBlock(
            callout_font,
            "JA: trust CER over WER/WCMR"
            "output inflates word-count metrics.",
        ),
        "\n",
        TextBlock(
            callout_font,
            "AR: Qwen3-ASR performance is poor and inconsistent for Arabic.",
        ),
    )
    ws.merge_cells(start_row=row, start_column=LANG_COL, end_row=row, end_column=merge_end)
    c = ws.cell(row=row, column=LANG_COL, value=note)
    c.font = FONT_LEGEND
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 36
    return row + 1


def _write_section_header(ws, row: int, text: str, n_cols: int,
                          start_col: int = 1) -> int:
    end_col = start_col + n_cols - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = FONT_SECTION
    c.fill = FILL_SECTION
    c.alignment = ALIGN_LEFT
    c.border = BORDER
    return row + 1


def _padded_end_col(content_end_col: int) -> int:
    return content_end_col + TRAILING_PAD_COLS


def _batch_sort_key(batch_name: str) -> tuple[int, int | str]:
    """Newest ``delivery_batch_MMDDYYYY`` first; other names sort after, A→Z."""
    prefix = "delivery_batch_"
    if batch_name.startswith(prefix):
        suffix = batch_name[len(prefix):]
        if suffix.isdigit():
            return (0, -int(suffix))
    return (1, batch_name)


def _cell_has_border(cell) -> bool:
    if isinstance(cell, MergedCell):
        return True
    b = cell.border
    return bool(b and any((b.left.style, b.right.style, b.top.style, b.bottom.style)))


def _style_trailing_pad(
    ws,
    pad_after_col: int,
    pad_through_col: int | None = None,
    *,
    first_row: int = 1,
    last_row: int | None = None,
) -> None:
    """Widen columns past the data block; touch only empty non-table cells for scroll."""
    end = pad_through_col if pad_through_col is not None else (
        pad_after_col + TRAILING_PAD_COLS
    )
    last_row = last_row if last_row is not None else ws.max_row
    for col in range(pad_after_col + 1, end + 1):
        ws.column_dimensions[get_column_letter(col)].width = TRAILING_PAD_WIDTH
        for row in range(first_row, last_row + 1):
            cell = ws.cell(row=row, column=col)
            if _cell_has_border(cell):
                continue
            if cell.value is None:
                cell.value = ""


def _apply_batch_sheet_scroll_pad(
    ws, first_row: int, last_row: int, content_end_col: int,
) -> None:
    """Four empty columns immediately after the rightmost data column."""
    _style_trailing_pad(
        ws,
        content_end_col,
        content_end_col + TRAILING_PAD_COLS,
        first_row=first_row,
        last_row=last_row,
    )


def _style_gap_column(ws, start_row: int, end_row: int, col: int = GAP_COL) -> None:
    """Narrow grey gutter between side-by-side tables."""
    ws.column_dimensions[get_column_letter(col)].width = 2
    fill = PatternFill("solid", fgColor="E7E6E6")
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).fill = fill


def _deter_cell_style(val, key: str) -> tuple[PatternFill | None, Font]:
    if key == "deter_pass_label":
        if val == "PASS" or (isinstance(val, str) and val.startswith("PASS")):
            return FILL_DETER_PASS_LABEL, FONT_DELTA_GOOD
        if val == "FAIL" or (isinstance(val, str) and val.startswith("FAIL")):
            return FILL_DETER_FAIL_LABEL, FONT_DELTA_BAD
        return None, FONT_BODY
    if isinstance(val, (int, float)) and key.startswith("SPK"):
        if val <= DETER_PASS_THRESHOLD_PCT:
            return FILL_DETER_PASS, FONT_DELTA_GOOD
        return FILL_DETER_FAIL, FONT_DELTA_BAD
    return None, FONT_BODY


def _write_table(ws, start_row: int, headers: list[tuple], rows: list[dict],
                 total_labels: set[str] | None = None,
                 start_col: int = 1) -> int:
    total_labels = total_labels or {"ALL"}

    for i, (label, _, _, _) in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=start_row, column=col, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    text_keys = {"session_id", "language", "batch", "deter_pass_label"}
    for i, record in enumerate(rows):
        r = start_row + 1 + i
        is_total = record.get("language") in total_labels
        is_alt = i % 2 == 1 and not is_total
        for j, (_, key, fmt, col_kind) in enumerate(headers):
            col = start_col + j
            val = record.get(key)
            if val is None and col_kind is True:
                display = "—"
            else:
                display = val
            cell = ws.cell(row=r, column=col, value=display)
            cell.border = BORDER

            if col_kind is True and val is not None:
                band = _delta_band(val)
                cell.fill = FILL_DELTA[band]
                cell.font = _delta_font(band)
                cell.number_format = fmt or "+0.0;-0.0"
                cell.alignment = ALIGN_CENTER
            elif col_kind == "deter":
                fill, font = _deter_cell_style(val, key)
                cell.font = font
                if fill is not None:
                    cell.fill = fill
                elif is_total:
                    cell.fill = FILL_TOTAL
                    cell.font = FONT_TOTAL
                elif is_alt:
                    cell.fill = FILL_ALT
                if fmt and val is not None:
                    cell.number_format = fmt
                cell.alignment = (
                    ALIGN_LEFT if key in text_keys else ALIGN_RIGHT
                )
            else:
                cell.font = FONT_TOTAL if is_total else FONT_BODY
                if key == "deter_pass_label" and val in ("PASS", "FAIL"):
                    fill, font = _deter_cell_style(val, key)
                    cell.font = font
                    if fill is not None:
                        cell.fill = fill
                elif is_total:
                    cell.fill = FILL_TOTAL
                elif is_alt:
                    cell.fill = FILL_ALT
                if fmt and val is not None:
                    cell.number_format = fmt
                cell.alignment = (
                    ALIGN_LEFT if key in text_keys else ALIGN_RIGHT
                )

    return start_row + 1 + len(rows)


def _write_top_errors_table(ws, start_row: int, headers: list[tuple],
                            rows: list[dict]) -> int:
    """Flat table for top erroneous segments; wrap long transcript columns."""
    for i, (label, _, _, _) in enumerate(headers):
        cell = ws.cell(row=start_row, column=1 + i, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for i, record in enumerate(rows):
        r = start_row + 1 + i
        is_alt = i % 2 == 1
        for j, (_, key, fmt, _) in enumerate(headers):
            col = 1 + j
            val = record.get(key)
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            cell.font = FONT_BODY
            if is_alt:
                cell.fill = FILL_ALT
            if key in TOP_ERRORS_WRAP_KEYS:
                cell.alignment = wrap
            elif key in TOP_ERRORS_TEXT_KEYS:
                cell.alignment = ALIGN_LEFT
            else:
                cell.alignment = ALIGN_RIGHT
            if fmt and val is not None:
                cell.number_format = fmt

    return start_row + 1 + len(rows)


def _autosize_columns(ws, col_ranges: list[tuple[int, int]],
                      min_width: int = 9, max_width: int = 22) -> None:
    for start, end in col_ranges:
        for col_idx in range(start, end + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(
                max(max_len + 2, min_width), max_width)


def _deter_block_cols(conv_headers: list[tuple]) -> tuple[int, int]:
    """Return (gap_col, deter_lang_col) immediately after the conv metrics table."""
    conv_end_col = CONV_COL + len(conv_headers) - 1
    gap_col = conv_end_col + 1
    return gap_col, gap_col + 1


def _write_metrics_sheet(ws, title: str, subtitle: str,
                         conv_headers: list[tuple], conv_records: list[dict],
                         lang_records: list[dict],
                         deter_conv_headers: list[tuple] | None = None,
                         deter_conv_records: list[dict] | None = None,
                         deter_lang_records: list[dict] | None = None) -> None:
    """Transcription tables (A→) and optional DetER tables after the conv block."""
    conv_end_col = CONV_COL + len(conv_headers) - 1
    deter_gap_col, deter_lang_col = _deter_block_cols(conv_headers)
    content_end_col = conv_end_col
    deter_conv_col: int | None = None
    if deter_conv_headers and deter_conv_records is not None:
        deter_lang_headers = build_deter_lang_headers(
            speaker_columns(deter_conv_records))
        deter_lang_width = len(deter_lang_headers)
        deter_mid_gap = deter_lang_col + deter_lang_width
        deter_conv_col = deter_mid_gap + 1
        content_end_col = max(
            content_end_col, deter_conv_col + len(deter_conv_headers) - 1)

    padded_end_col = content_end_col + TRAILING_PAD_COLS
    row = _write_title_block(ws, title, subtitle, padded_end_col)
    row = _write_reading_guide(ws, row, padded_end_col)

    table_row = row
    row = _write_section_header(
        ws, row,
        "Summary by language (micro-averaged)",
        len(LANG_HEADERS), start_col=LANG_COL,
    )
    row = _write_section_header(
        ws, table_row,
        f"Per conversation ({len(conv_records)} session(s))",
        len(conv_headers), start_col=CONV_COL,
    )
    if deter_conv_col is not None and deter_lang_records is not None:
        deter_lang_headers = build_deter_lang_headers(
            speaker_columns(deter_conv_records))
        _write_section_header(
            ws, table_row,
            "DetER by language (avg per channel)",
            len(deter_lang_headers), start_col=deter_lang_col,
        )
        _write_section_header(
            ws, table_row,
            f"DetER per conversation ({len(deter_conv_records)} session(s))",
            len(deter_conv_headers), start_col=deter_conv_col,
        )
    row = max(row, table_row + 1)

    header_row = row
    lang_end = _write_table(
        ws, row, LANG_HEADERS, lang_records, start_col=LANG_COL)
    conv_end = _write_table(
        ws, row, conv_headers, conv_records, start_col=CONV_COL)
    data_end = max(lang_end, conv_end)

    if deter_conv_col is not None and deter_conv_records is not None:
        deter_lang_headers = build_deter_lang_headers(
            speaker_columns(deter_conv_records))
        deter_lang_end = _write_table(
            ws, row, deter_lang_headers, deter_lang_records or [],
            start_col=deter_lang_col)
        deter_conv_end = _write_table(
            ws, row, deter_conv_headers, deter_conv_records,
            start_col=deter_conv_col)
        data_end = max(data_end, deter_lang_end, deter_conv_end)
        deter_mid_gap = deter_lang_col + len(deter_lang_headers)
        _style_gap_column(ws, table_row, data_end - 1, col=deter_gap_col)
        _style_gap_column(ws, table_row, data_end - 1, col=deter_mid_gap)

    _style_gap_column(ws, table_row, data_end - 1)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=LANG_COL)
    col_ranges = [
        (LANG_COL, LANG_COL + len(LANG_HEADERS) - 1),
        (CONV_COL, CONV_COL + len(conv_headers) - 1),
    ]
    if deter_conv_col is not None and deter_conv_headers is not None:
        col_ranges.append((deter_lang_col, deter_lang_col + len(deter_lang_headers) - 1))
        col_ranges.append((deter_conv_col, deter_conv_col + len(deter_conv_headers) - 1))
    _autosize_columns(ws, col_ranges)
    _apply_batch_sheet_scroll_pad(ws, table_row, data_end - 1, content_end_col)


def write_batch_sheet(ws, batch_name: str, records: list[dict], generated: str,
                      deter_records: list[dict] | None = None) -> None:
    title = f"Batch — {batch_name}"
    subtitle = (
        f"Human-annotated transcript vs Qwen3-ASR  ·  vs baseline reference  ·  "
        f"{len(records)} conversation(s)  ·  Generated {generated} UTC"
    )
    deter_kwargs: dict = {}
    if deter_records:
        speaker_cols = speaker_columns(deter_records)
        deter_kwargs = {
            "deter_conv_headers": build_deter_conv_headers(speaker_cols),
            "deter_conv_records": deter_records,
            "deter_lang_records": deter_language_summary(deter_records, speaker_cols),
        }
    _write_metrics_sheet(
        ws, title, subtitle,
        CONV_HEADERS, records, language_summary(records),
        **deter_kwargs,
    )


def write_all_batches_sheet(ws, all_records: list[dict], generated: str,
                            all_deter_records: list[dict] | None = None) -> None:
    conv_headers = [("Batch", "batch", None, False)] + list(CONV_HEADERS)
    title = "All batches — combined"
    subtitle = (
        f"Human-annotated transcript vs Qwen3-ASR  ·  vs baseline reference  ·  "
        f"{len(all_records)} conversation(s)  ·  Generated {generated} UTC"
    )
    deter_kwargs: dict = {}
    if all_deter_records:
        speaker_cols = speaker_columns(all_deter_records)
        deter_kwargs = {
            "deter_conv_headers": build_deter_conv_headers(
                speaker_cols, include_batch=True),
            "deter_conv_records": all_deter_records,
            "deter_lang_records": deter_language_summary(
                all_deter_records, speaker_cols),
        }
    _write_metrics_sheet(
        ws, title, subtitle,
        conv_headers, all_records, language_summary(all_records),
        **deter_kwargs,
    )


def write_reference_sheet(ws) -> None:
    """Baseline reference corpus — separate tab for lookup."""
    ref_end = len(REF_HEADERS)
    deter_end = REF_DETER_COL + len(REF_DETER_HEADERS) - 1
    content_end_col = max(ref_end, deter_end)
    padded_end_col = _padded_end_col(content_end_col)
    row = _write_title_block(
        ws,
        "Baseline reference",
        "Independent vendor corpus · identical scoring pipeline (Qwen3-ASR, normalization, metrics)",
        padded_end_col,
    )
    row += 1
    row = _write_section_header(ws, row, "Per language", ref_end)
    ref_rows = []
    for lang in sorted(k for k in REFERENCE if k != "ALL"):
        ref = REFERENCE[lang]
        ref_rows.append({
            "language": lang,
            "n_scored": REF_N_SCORED[lang],
            "wer": ref["wer"],
            "cer": ref["cer"],
            "wcmr": ref["wcmr"],
            "gt3_pct": ref["gt3_pct"],
            "deter_pass_max": DETER_PASS_THRESHOLD_PCT,
        })
    all_ref = REFERENCE["ALL"]
    ref_rows.append({
        "language": "ALL",
        "n_scored": REF_N_SCORED["ALL"],
        "wer": all_ref["wer"],
        "cer": all_ref["cer"],
        "wcmr": all_ref["wcmr"],
        "gt3_pct": all_ref["gt3_pct"],
        "deter_pass_max": DETER_PASS_THRESHOLD_PCT,
    })
    table_start = row
    _write_table(ws, row, REF_HEADERS, ref_rows)
    _write_table(ws, row, REF_DETER_HEADERS, ref_rows, start_col=REF_DETER_COL)
    data_end = table_start + len(ref_rows)
    _style_gap_column(ws, table_start - 1, data_end, col=ref_end + 1)
    _autosize_columns(ws, [(1, ref_end), (REF_DETER_COL, deter_end)])
    _style_trailing_pad(ws, content_end_col)


def write_top_erroneous_segments_sheet(ws, rows: list[dict], generated: str) -> None:
    """Worst segments per speaker — sourced from SPK*_top_errors.json."""
    content_end_col = len(TOP_ERRORS_HEADERS)
    padded_end_col = _padded_end_col(content_end_col)
    row = _write_title_block(
        ws,
        "Top erroneous segments",
        f"Per-speaker worst segments (rank_error_segments.py)  ·  "
        f"{len(rows)} row(s)  ·  Generated {generated} UTC",
        padded_end_col,
    )
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=padded_end_col)
    note = ws.cell(
        row=row, column=1,
        value=(
            "Japanese and Korean rows ranked by CER (character errors); all other "
            "languages by WER (word errors). Error rate % can exceed 100 when "
            "insertions dominate. Use Start/End to locate audio in SPK*.wav."
        ),
    )
    note.font = FONT_LEGEND
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 28
    row += 1

    header_row = row
    _write_top_errors_table(ws, row, TOP_ERRORS_HEADERS, rows)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    col_widths = {
        "batch": 24,
        "session_id": 22,
        "speaker": 8,
        "language": 8,
        "rank_metric": 10,
        "ref_norm": 44,
        "hyp_norm": 44,
        "ref_raw": 44,
        "hyp_raw": 44,
    }
    for i, (_, key, _, _) in enumerate(TOP_ERRORS_HEADERS):
        letter = get_column_letter(1 + i)
        ws.column_dimensions[letter].width = col_widths.get(key, 11)
    _style_trailing_pad(ws, content_end_col)


def write_definitions_sheet(ws) -> None:
    """Glossary of metric columns used in the batch / All Batches tabs."""
    content_end_col = 2
    padded_end_col = _padded_end_col(content_end_col)
    row = _write_title_block(
        ws,
        "Metric definitions",
        "Column glossary · human-annotated transcript vs Qwen3-ASR-1.7B",
        padded_end_col,
    )
    row += 1
    row = _write_section_header(ws, row, "Columns", content_end_col)

    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for col, label in enumerate(("Metric", "Definition"), start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    for i, (metric, definition) in enumerate(METRIC_DEFINITIONS):
        r = row + 1 + i
        m_cell = ws.cell(row=r, column=1, value=metric)
        m_cell.font = FONT_BODY
        m_cell.border = BORDER
        m_cell.alignment = ALIGN_LEFT
        d_cell = ws.cell(row=r, column=2, value=definition)
        d_cell.font = FONT_BODY
        d_cell.border = BORDER
        d_cell.alignment = wrap
        if i % 2 == 1:
            m_cell.fill = FILL_ALT
            d_cell.fill = FILL_ALT
        ws.row_dimensions[r].height = 36

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 88
    freeze_row = row + 1

    if TOP_ERRORS_DEFINITIONS:
        row = row + 1 + len(METRIC_DEFINITIONS) + 1
        row = _write_section_header(ws, row, "Top erroneous segments tab", content_end_col)
        for col, label in enumerate(("Column", "Definition"), start=1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER
        for i, (metric, definition) in enumerate(TOP_ERRORS_DEFINITIONS):
            r = row + 1 + i
            m_cell = ws.cell(row=r, column=1, value=metric)
            m_cell.font = FONT_BODY
            m_cell.border = BORDER
            m_cell.alignment = ALIGN_LEFT
            d_cell = ws.cell(row=r, column=2, value=definition)
            d_cell.font = FONT_BODY
            d_cell.border = BORDER
            d_cell.alignment = wrap
            if i % 2 == 1:
                m_cell.fill = FILL_ALT
                d_cell.fill = FILL_ALT
            ws.row_dimensions[r].height = 36

    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    _style_trailing_pad(ws, content_end_col)


def build_workbook(batches: dict[str, list[dict]], top_error_rows: list[dict],
                   generated: str,
                   deter_batches: dict[str, list[dict]] | None = None) -> Workbook:
    deter_batches = deter_batches or {}
    wb = Workbook()
    wb.remove(wb.active)

    sheet_idx = 0
    ws_def = wb.create_sheet(title=_safe_sheet_name("Definitions"), index=sheet_idx)
    write_definitions_sheet(ws_def)
    sheet_idx += 1

    ws_ref = wb.create_sheet(title=_safe_sheet_name("Reference"), index=sheet_idx)
    write_reference_sheet(ws_ref)
    sheet_idx += 1

    all_records: list[dict] = []
    all_deter_records: list[dict] = []
    for batch_name, records in batches.items():
        all_records.extend(records)
        deter_records = deter_batches.get(batch_name)
        if deter_records:
            all_deter_records.extend(deter_records)

    if all_records:
        all_records.sort(key=lambda r: (r["batch"], r["session_id"]))
        all_deter = all_deter_records or None
        if all_deter:
            all_deter.sort(key=lambda r: (r["batch"], r["session_id"]))
        ws_all = wb.create_sheet(
            title=_safe_sheet_name("All Batches"), index=sheet_idx)
        write_all_batches_sheet(ws_all, all_records, generated, all_deter)
        sheet_idx += 1

    for batch_name, records in sorted(
        batches.items(), key=lambda item: _batch_sort_key(item[0]),
    ):
        deter_records = deter_batches.get(batch_name)
        ws = wb.create_sheet(title=_safe_sheet_name(batch_name), index=sheet_idx)
        write_batch_sheet(ws, batch_name, records, generated, deter_records)
        sheet_idx += 1

    if top_error_rows:
        ws_top = wb.create_sheet(
            title=_safe_sheet_name("top_erroneous_segments"), index=sheet_idx)
        write_top_erroneous_segments_sheet(ws_top, top_error_rows, generated)

    return wb


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    add_scope_args(parser, with_file=False)
    parser.add_argument(
        "--output", "-o", default="reports/transcription_accuracy_metrics.xlsx",
        help="Output .xlsx path (default: reports/transcription_accuracy_metrics.xlsx).",
    )
    args = parser.parse_args(argv)

    root = Path(args.conversations)
    try:
        batches = discover_records(root, args.batch)
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}")
        return 1

    if not batches:
        print("No metrics.json files found in scope.")
        return 1

    deter_batches = discover_deter_records(root, args.batch)
    top_error_rows = discover_top_error_segments(root, args.batch)

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    wb = build_workbook(batches, top_error_rows, generated, deter_batches)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    n_conv = sum(len(v) for v in batches.values())
    n_deter = sum(len(v) for v in deter_batches.values())
    deter_note = (
        f" · DetER on {n_deter} conversation(s)"
        if n_deter else " · no deter.json found"
    )
    top_note = (
        f" + top_erroneous_segments ({len(top_error_rows)} row(s))"
        if top_error_rows else " (no SPK*_top_errors.json found)"
    )
    print(f"Wrote {out_path.resolve()}")
    print(
        f"  Tabs: Definitions, Reference, All Batches, "
        f"{len(batches)} batch tab(s) (newest first)"
        f"{top_note}  ({n_conv} conversation(s){deter_note})."
    )
    for batch_name, records in sorted(
        batches.items(), key=lambda item: _batch_sort_key(item[0]),
    ):
        langs = sorted({r["language"] for r in records})
        print(f"    {batch_name}: {len(records)} conversation(s)  [{', '.join(langs)}]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
